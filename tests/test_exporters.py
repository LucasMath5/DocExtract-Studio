"""Tests for CSV and Excel export files."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from pdf_extractor.exporters.base import (
    ExportDataset,
    build_batch_export_dataset,
    build_export_dataset,
)
from pdf_extractor.exporters.csv_exporter import CsvExporter
from pdf_extractor.exporters.excel_exporter import ExcelExporter
from pdf_extractor.models.extraction_field import ExtractionField
from pdf_extractor.models.extraction_result import (
    ExtractionMethod,
    ExtractionResult,
    ExtractionStatus,
)
from pdf_extractor.models.field_region import FieldRegion
from pdf_extractor.models.batch_result import (
    BatchDocumentResult,
    BatchDocumentStatus,
)


def sample_dataset() -> ExportDataset:
    """Build an ordered dataset containing accents and an empty field."""
    fields = (
        ExtractionField("cliente", "Cliente", FieldRegion(0, 1, 1, 10, 10)),
        ExtractionField("observacao", "Observação", FieldRegion(0, 1, 20, 10, 10)),
    )
    results = (
        ExtractionResult(
            "cliente",
            "Cliente",
            0,
            "Empresa São João",
            ExtractionStatus.SUCCESS,
        ),
        ExtractionResult(
            "observacao",
            "Observação",
            0,
            "",
            ExtractionStatus.EMPTY,
        ),
    )
    return build_export_dataset("documento.pdf", fields, results)


def test_csv_export_preserves_order_accents_and_empty_values(tmp_path: Path) -> None:
    """CSV should contain one ordered document row in UTF-8."""
    output = tmp_path / "dados.csv"

    CsvExporter().export(output, sample_dataset())

    with output.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows == [
        ["arquivo", "Cliente", "Observação"],
        ["documento.pdf", "Empresa São João", ""],
    ]


def test_excel_export_creates_readable_formatted_workbook(tmp_path: Path) -> None:
    """XLSX should contain one sheet, ordered values, and styled headers."""
    output = tmp_path / "dados.xlsx"

    ExcelExporter().export(output, sample_dataset())

    workbook = load_workbook(output)
    worksheet = workbook["Dados extraídos"]
    assert list(worksheet.values) == [
        ("arquivo", "Cliente", "Observação"),
        ("documento.pdf", "Empresa São João", None),
    ]
    assert worksheet.freeze_panes == "A2"
    assert worksheet["A1"].font.bold is True
    assert worksheet.auto_filter.ref == "A1:C2"
    workbook.close()


def test_batch_exporters_write_one_consolidated_row_per_document(
    tmp_path: Path,
) -> None:
    """CSV and XLSX should include statuses, errors, and every processed PDF."""
    field = ExtractionField(
        "cliente",
        "Cliente",
        FieldRegion(0, 1, 1, 10, 10),
    )
    success_result = ExtractionResult(
        "cliente",
        "Cliente",
        0,
        "Empresa Exemplo",
        ExtractionStatus.SUCCESS,
        method=ExtractionMethod.NATIVE_TEXT,
    )
    documents = (
        BatchDocumentResult(
            tmp_path / "sucesso.pdf",
            (success_result,),
            BatchDocumentStatus.SUCCESS,
        ),
        BatchDocumentResult(
            tmp_path / "falha.pdf",
            (),
            BatchDocumentStatus.FAILURE,
            "PDF inválido",
        ),
    )
    dataset = build_batch_export_dataset((field,), documents)
    csv_path = tmp_path / "lote.csv"
    excel_path = tmp_path / "lote.xlsx"

    CsvExporter().export(csv_path, dataset)
    ExcelExporter().export(excel_path, dataset)

    with csv_path.open(encoding="utf-8-sig", newline="") as stream:
        assert list(csv.reader(stream)) == [
            ["arquivo", "status", "método", "erro", "Cliente"],
            [
                "sucesso.pdf",
                "sucesso",
                "texto nativo",
                "",
                "Empresa Exemplo",
            ],
            ["falha.pdf", "falha", "", "PDF inválido", ""],
        ]
    workbook = load_workbook(excel_path)
    worksheet = workbook["Dados extraídos"]
    assert list(worksheet.values) == [
        ("arquivo", "status", "método", "erro", "Cliente"),
        ("sucesso.pdf", "sucesso", "texto nativo", None, "Empresa Exemplo"),
        ("falha.pdf", "falha", None, "PDF inválido", None),
    ]
    assert worksheet.auto_filter.ref == "A1:E3"
    workbook.close()
