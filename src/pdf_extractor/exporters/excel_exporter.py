"""Excel exporter for one extracted PDF document."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pdf_extractor.exporters.base import ExportDataset, ExportError


class ExcelExporter:
    """Write an export dataset to a compact, readable XLSX workbook."""

    def export(self, file_path: Path, dataset: ExportDataset) -> None:
        """Create one worksheet containing headers and one document row."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Dados extraídos"
        worksheet.append(dataset.headers)
        worksheet.append(dataset.values)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        for column_index, header in enumerate(dataset.headers, start=1):
            value = dataset.values[column_index - 1]
            width = min(max(len(header), len(value)) + 2, 50)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        try:
            workbook.save(file_path)
        except OSError as error:
            raise ExportError("Não foi possível salvar o arquivo Excel.") from error
        finally:
            workbook.close()
